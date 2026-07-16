"""Migrador SLEP - pipeline completo de procesamiento de licencias médicas.

Este módulo concentra la lógica de punta a punta del migrador. Las reglas de
negocio declarativas (regex y mapas canónicos) viven en :mod:`slep.constants`
y la normalización de bajo nivel en :mod:`slep.utils`; aquí se orquestan.

Etapas del pipeline (ver :func:`procesar`):

1. **Lectura** (:func:`leer_fuente`, :func:`leer_dim_establecimiento`):
   carga la planilla madre (hoja ``DATOS`` de funcionarios, hojas ``LM*`` de
   hechos) y la tabla maestra de establecimientos.
2. **Clasificación / canonización** (:func:`clasificar_generico`,
   :func:`clasificar_resolucion`, :func:`clasificar_afp`,
   :func:`clasificar_establecimiento`): convierte el texto libre histórico a
   los valores canónicos oficiales, marcando lo que requiere revisión.
3. **Migración** (:func:`migrar_hechos`, :func:`migrar_descuentos`,
   :func:`construir_dim_afp`): consolida los hechos deduplicados, extrae los
   descuentos por período y deriva la dimensión de AFPs con sus tasas.
4. **Escritura** (``escribir_*``): genera los cinco archivos del modelo
   estrella (tres dimensiones + dos tablas de hechos) con formato,
   validaciones de datos y fórmulas de autorrelleno para la imputación.

Salidas (dict ``{nombre: bytes}`` devuelto por :func:`procesar`):
    ``01_Dim_Funcionario.xlsx``, ``02_Dim_Establecimiento.xlsx``,
    ``03_Dim_AFP.xlsx``, ``04_Hechos_Licencias.xlsx``,
    ``05_Hechos_Descuentos.xlsx`` y ``SLEP_files.zip`` (más
    ``Dashboard_Licencias.pbit`` si se entrega la plantilla).

Los comentarios ``RB-*`` referencian el catálogo de reglas de negocio del
documento técnico (``docs/Documento_Tecnico_Migrador_SLEP.md``).

Uso:
    from slep.core import procesar
    resultados = procesar(licencias_bytes, establecimientos_bytes)

Nota de despliegue:
    El paquete está pensado para ejecutarse en el navegador bajo
    Pyodide/WebAssembly (ver ``assets/worker.js``): opera sobre ``bytes`` en
    memoria y evita cualquier acceso a disco o red.
"""

from __future__ import annotations

import io
import re
import zipfile
from collections import Counter
from datetime import datetime, date
from difflib import get_close_matches

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table

from .constants import RE_TIPO, RE_INST, RE_RESOL, RE_ESTABLECIMIENTO, AFP_MAP, MESES_MAP
from .utils import norm, norm_rut, _is_excel_error

# CLASIFICADORES

def _match_canon(n: str, patterns: dict):
    """Busca el primer valor canónico cuyas regex coincidan con el texto.

    Recorre ``patterns`` en orden de inserción (la primera coincidencia
    gana), probando cada patrón con ``re.search`` (coincidencia parcial,
    no anclada al string completo).

    Args:
        n: Texto **ya normalizado** con :func:`slep.utils.norm` (RB-01).
        patterns: Dict ``{valor_canonico: [regex, ...]}`` de ``constants``.

    Returns:
        El valor canónico de la primera clave que coincide, o ``None`` si
        ninguna lo hace.
    """
    for canon, pats in patterns.items():
        if any(re.search(p, n) for p in pats):
            return canon
    return None


def extraer_valores_unicos(hojas_hechos: list, *alt_names: str) -> list:
    """Extrae los valores únicos no vacíos de una columna en todas las hojas de hechos.

    Acepta varios nombres alternativos de columna (``alt_names``) y usa el
    primero que exista en cada hoja, porque las hojas LM de distintos años
    varían en tildes y nomenclatura (ej. "Institución Salud" /
    "Institucion Salud").

    Args:
        hojas_hechos: Lista de tuplas ``(nombre_hoja, worksheet, fila_header,
            headers)`` producida por :func:`leer_fuente`.
        *alt_names: Nombres alternativos de la columna buscada.

    Returns:
        Lista ordenada alfabéticamente con los valores crudos únicos
        (como ``str`` y sin espacios extremos).
    """
    vals = set()
    for _, ws, hrow, headers in hojas_hechos:
        for n in alt_names:
            if n not in headers:
                continue
            idx = headers.index(n)
            for row in ws.iter_rows(min_row=hrow + 1, values_only=True):
                v = row[idx]
                if v is not None and str(v).strip():
                    vals.add(str(v).strip())
            break
    return sorted(vals)


def clasificar_generico(raw, canon_list: list, patterns: dict, cutoff: float = 0.6):
    """Clasifica texto libre en cascada: regex -> fuzzy -> revisión (RB-04).

    Estrategia de negocio:
        1. Error de Excel -> vacío (RB-03).
        2. Vacío -> vacío.
        3. Coincidencia por regex sobre el texto normalizado -> ``"OK"``.
        4. Sin regex: coincidencia difusa (:func:`difflib.get_close_matches`)
           contra la lista canónica, con umbral ``cutoff`` ->
           ``"Corregido (revisar)"``. Cubre errores ortográficos leves.
        5. Sin nada: se conserva el valor crudo y se marca
           ``"REVISAR: no reconocido"``.

    Args:
        raw: Valor crudo de la celda.
        canon_list: Lista de valores canónicos admitidos.
        patterns: Dict de regex de ``constants`` asociado a la dimensión.
        cutoff: Umbral de similitud (0-1) para la corrección difusa.

    Returns:
        Tupla ``(valor_canonico, estado)``. ``valor_canonico`` es ``None``
        cuando el dato es vacío/error; ``estado`` es uno de ``"OK"``,
        ``"Vacio"``, ``"Vacio (error Excel)"``, ``"Corregido (revisar)"`` o
        ``"REVISAR: no reconocido"``.
    """
    if _is_excel_error(raw):
        return None, "Vacio (error Excel)"
    n = norm(raw)
    if not n:
        return None, "Vacio"
    canon = _match_canon(n, patterns)
    if canon:
        return canon, "OK"
    match = get_close_matches(n, [norm(x) for x in canon_list], n=1, cutoff=cutoff)
    if match:
        idx = [norm(x) for x in canon_list].index(match[0])
        return canon_list[idx], "Corregido (revisar)"
    return raw, "REVISAR: no reconocido"


def clasificar_resolucion(raw, canon_list: list):
    """Clasifica la Resolución Médica con regla adicional para datos legacy (RB-05).

    Igual que :func:`clasificar_generico` sobre el catálogo ``RE_RESOL``,
    pero antes descarta el caso histórico en que la columna "Resolución
    Médica" contiene el **número** de la resolución (ej. "12345678-9",
    "1.234.567") en lugar del estado: esos valores no son un estado y se
    reportan como legacy sin marcar inconsistencia.

    Args:
        raw: Valor crudo de la celda.
        canon_list: Lista canónica de estados (claves de ``RE_RESOL``).

    Returns:
        Tupla ``(valor_canonico, estado)``; además de los estados de
        :func:`clasificar_generico`, puede devolver
        ``(None, "LEGACY: es un N de resolucion, no un estado")``.
    """
    if _is_excel_error(raw):
        return None, "Vacio (error Excel)"
    n = norm(raw)
    if not n:
        return None, "Vacio"
    if re.fullmatch(r"[\d\.\-]+", n.replace(" ", "")):
        return None, "LEGACY: es un N de resolucion, no un estado"
    canon = _match_canon(n, RE_RESOL)
    if canon:
        return canon, "OK"
    match = get_close_matches(n, [norm(x) for x in canon_list], n=1, cutoff=0.6)
    if match:
        idx = [norm(x) for x in canon_list].index(match[0])
        return canon_list[idx], "Corregido (revisar)"
    return raw, "REVISAR: no reconocido"


def clasificar_afp(raw):
    """Clasifica la AFP y extrae su tasa de cotización (RB-06 y RB-07).

    La planilla madre registra la AFP como texto libre, frecuentemente con
    la tasa entre paréntesis: ``"Habitat (11,27)"``, ``"provida"``, etc.
    La regex ``^([^\\(]+?)\\s*(?:\\(([\\d.,]+)\\))?\\s*$`` separa el nombre
    de la tasa opcional, admitiendo coma o punto decimal.

    La canonización del nombre es por **igualdad exacta** sobre el texto
    normalizado contra las claves de ``AFP_MAP``; si falla, se intenta
    corrección difusa (cutoff 0.6).

    Args:
        raw: Valor crudo de la celda ``A.F.P.``.

    Returns:
        Tupla ``(afp_canonica, tasa, estado)`` donde ``tasa`` es ``float`` o
        ``None`` si la celda no la traía; ``estado`` es ``"OK"``, ``"Vacio"``,
        ``"Vacio (error Excel)"``, ``"Corregido (revisar)"`` o
        ``"REVISAR: AFP no reconocida"``.
    """
    if not raw or not str(raw).strip():
        return None, None, "Vacio"
    if _is_excel_error(raw):
        return None, None, "Vacio (error Excel)"
    m = re.match(r"^([^\(]+?)\s*(?:\(([\d.,]+)\))?\s*$", str(raw).strip())
    if not m:
        return raw, None, "REVISAR: no reconocido"
    nombre_n = norm(m.group(1))
    canon = AFP_MAP.get(nombre_n)
    estado = "OK"
    if canon is None:
        match = get_close_matches(nombre_n, list(AFP_MAP.keys()), n=1, cutoff=0.6)
        if not match:
            return raw, None, "REVISAR: AFP no reconocida"
        canon, estado = AFP_MAP[match[0]], "Corregido (revisar)"
    tasa = None
    if m.group(2):
        try:
            tasa = float(m.group(2).replace(",", "."))
        except ValueError:
            pass
    return canon, tasa, estado


def clasificar_establecimiento(raw, catalogo_norm: dict, catalogo_patterns: dict):
    """Clasifica el establecimiento en cascada de 4 niveles (RB-08).

    A diferencia de :func:`clasificar_generico`, aquí existe una **tabla
    maestra oficial** (``Establecimientos.xlsx``), por lo que la cascada es:

        1. **Exacto**: el texto normalizado coincide con un nombre del
           catálogo maestro -> ``"OK"``.
        2. **Regex**: coincide con un patrón de ``RE_ESTABLECIMIENTO``
           (variantes ortográficas conocidas) -> ``"OK (regex)"``.
        3. **Fuzzy**: similitud difusa con el catálogo, umbral alto (0.82,
           más estricto que el genérico para no fusionar escuelas
           parecidas) -> ``"Corregido (revisar)"``.
        4. **Nuevo**: no existe en el maestro -> se conserva el crudo, se
           marca ``"NUEVO: ..."`` y el pipeline lo agrega a
           ``Dim_Establecimiento`` con tipo "Otro" para revisión.

    Args:
        raw: Valor crudo de la celda de establecimiento/unidad.
        catalogo_norm: Dict ``{nombre_normalizado: nombre_canonico}`` del
            maestro de establecimientos.
        catalogo_patterns: Dict de regex (``RE_ESTABLECIMIENTO``).

    Returns:
        Tupla ``(establecimiento_canonico, estado)`` con estados ``"OK"``,
        ``"OK (regex)"``, ``"Corregido (revisar)"``, ``"Vacio"``,
        ``"Vacio (error Excel)"`` o ``"NUEVO: no esta en Dim_Establecimiento,
        se agrega y debe revisarse"``.
    """
    if _is_excel_error(raw):
        return None, "Vacio (error Excel)"
    n = norm(raw)
    if not n:
        return None, "Vacio"
    if n in catalogo_norm:
        return catalogo_norm[n], "OK"
    canon = _match_canon(n, catalogo_patterns)
    if canon:
        return canon, "OK (regex)"
    match = get_close_matches(n, list(catalogo_norm.keys()), n=1, cutoff=0.82)
    if match:
        return catalogo_norm[match[0]], "Corregido (revisar)"
    return raw, "NUEVO: no esta en Dim_Establecimiento, se agrega y debe revisarse"


def generar_listas_canonicas(hojas_hechos: list):
    """Deriva las listas canónicas y el log de clasificación desde el histórico.

    Recorre los valores crudos únicos de las tres dimensiones de texto libre
    (Tipo de Licencia, Institución de Salud y Resolución Médica), los
    clasifica y construye, para cada dimensión, la lista canónica ordenada
    (las claves oficiales de ``constants``) más un mapa de auditoría
    ``{canonico: [valores_crudos_que_mapearon]}`` que :func:`procesar`
    vuelca al log. Lo que no clasifica queda bajo la clave
    ``"_sin_clasificar"``.

    Args:
        hojas_hechos: Lista de tuplas ``(nombre, worksheet, fila_header,
            headers)`` de :func:`leer_fuente`.

    Returns:
        Tupla de tres elementos ``(tipo, institucion, resolucion)``, cada uno
        de la forma ``(lista_canonica, clasificacion)`` donde
        ``clasificacion`` es ``{canonico: [crudos]}``.
    """

    def _extraer(alt_names):
        vals = set()
        for _, ws, hrow, headers in hojas_hechos:
            for n in alt_names:
                if n not in headers:
                    continue
                idx = headers.index(n)
                for row in ws.iter_rows(min_row=hrow + 1, values_only=True):
                    v = row[idx]
                    if v is not None and str(v).strip():
                        vals.add(str(v).strip())
                break
        return vals

    # Tipo Licencia
    raw_tipo = _extraer(("Tipo Licencia",))
    tipo_clasif = {}
    for r in raw_tipo:
        canon, _ = clasificar_generico(r, list(RE_TIPO.keys()), RE_TIPO)
        if canon is None:
            continue
        if canon in RE_TIPO:
            tipo_clasif.setdefault(canon, []).append(r)
        else:
            tipo_clasif.setdefault("_sin_clasificar", []).append(r)
    tipo_canon = sorted(RE_TIPO.keys())

    # Institución Salud
    raw_inst = _extraer(("Institución Salud", "Institucion Salud"))
    inst_clasif = {}
    for r in raw_inst:
        canon, _ = clasificar_generico(r, list(RE_INST.keys()), RE_INST)
        if canon is None:
            continue
        if canon in RE_INST:
            inst_clasif.setdefault(canon, []).append(r)
        else:
            inst_clasif.setdefault("_sin_clasificar", []).append(r)
    inst_canon = sorted(RE_INST.keys())

    # Resolución Médica
    raw_resol = _extraer(("Resolución Médica", "Resolucion Medica", "Estado"))
    resol_clasif = {}
    for r in raw_resol:
        canon, _ = clasificar_resolucion(r, list(RE_RESOL.keys()))
        if canon is None:
            continue
        if canon in RE_RESOL:
            resol_clasif.setdefault(canon, []).append(r)
        else:
            resol_clasif.setdefault("_sin_clasificar", []).append(r)
    resol_canon = sorted(RE_RESOL.keys())

    return (tipo_canon, tipo_clasif), (inst_canon, inst_clasif), (resol_canon, resol_clasif)


# EXTRACTOR DE MONTOS DOBLES (Sistema / Pagado)

def extraer_montos_dobles(headers: list, row: tuple) -> dict:
    """Extrae los dos bloques de montos (Sistema vs. Pagado) de una fila (RB-11).

    La planilla madre lleva **dos circuitos de montos en paralelo** para cada
    licencia: el calculado por el sistema de remuneraciones (Dep/Netcore) y
    el efectivamente pagado por FONASA/ISAPRE. La estructura de columnas
    cambia según el año de la hoja:

    * **Patrón 2024**: ambos bloques repiten los mismos encabezados
      ("Monto de Subsidio", "Monto cotizacion previsional",
      "Monto Previsional Salud", "Total"); se distingue el bloque por la
      **primera o segunda ocurrencia** del encabezado.
    * **Patrón 2025/2026**: el bloque Sistema usa columnas llamadas
      exactamente ``AFP`` y ``SALUD`` (más "Monto de Subsidio" y "Total");
      el bloque Pagado mantiene los nombres largos, y su total puede
      llamarse "Total" o "Total Recuperado".

    La discriminante es la presencia simultánea de columnas ``afp`` y
    ``salud`` (en minúsculas, sin puntos) entre los encabezados.

    Args:
        headers: Encabezados de la hoja (valores crudos de la fila de
            encabezado).
        row: Tupla de valores de la fila de datos (misma longitud relativa
            que ``headers``).

    Returns:
        Dict con las 8 claves ``monto_subsidio_sistema``,
        ``monto_cotizacion_previsional_sistema``,
        ``monto_previsional_salud_sistema``, ``total_sistema``,
        ``monto_subsidio_pagado``, ``monto_cotizacion_previsional_pagado``,
        ``monto_previsional_salud_pagado`` y ``total_pagado``. Cada valor es
        ``float`` cuando la celda es numérica, el valor crudo si no se pudo
        convertir, o ``None`` si la columna no existe o la celda está vacía.
    """
    h_norm = [str(h).strip().lower() if h else "" for h in headers]

    def find_all(substr):
        return [i for i, h in enumerate(h_norm) if substr in h]

    def find_first(substr):
        idxs = find_all(substr)
        return idxs[0] if idxs else None

    def find_second(substr):
        idxs = find_all(substr)
        return idxs[1] if len(idxs) > 1 else None

    def val(idx):
        if idx is None or idx >= len(row):
            return None
        v = row[idx]
        if v is None:
            return None
        try:
            return float(v)
        except (ValueError, TypeError):
            return v

    # LM-2025/2026 usan 'AFP' y 'SALUD' como columnas de montos en el primer bloque,
    # mientras que LM-2024 usa 'cotizacion previsional' directamente.
    # Buscamos una columna exactamente llamada 'afp' (sin puntos) y 'salud'.
    tiene_afp_monto = any(h == "afp" for h in h_norm)
    tiene_salud_monto = any(h == "salud" for h in h_norm)

    if tiene_afp_monto and tiene_salud_monto:
        # Patrón 2025 / 2026
        # Primer bloque (Sistema): Monto de Subsidio | AFP | AFC | SALUD | Total
        # Segundo bloque (Pagado): Monto de Subsidio | Monto cotizacion previsional | Monto Previsional Salud | Total / Total Recuperado
        idx_sub_sist = find_first("monto de subsidio")
        idx_afp      = next((i for i, h in enumerate(h_norm) if h == "afp"), None)
        idx_salud    = next((i for i, h in enumerate(h_norm) if h == "salud"), None)
        idx_total_sist = find_first("total")

        idx_sub_pag  = find_second("monto de subsidio")
        idx_cot_pag  = find_first("cotizacion previsional")
        idx_prev_pag = find_first("monto previsional salud")
        # El total del segundo bloque puede llamarse 'Total' o 'Total Recuperado'
        idx_total_pag = find_first("total recuperado") or find_second("total")

        return {
            "monto_subsidio_sistema": val(idx_sub_sist),
            "monto_cotizacion_previsional_sistema": val(idx_afp),
            "monto_previsional_salud_sistema": val(idx_salud),
            "total_sistema": val(idx_total_sist),
            "monto_subsidio_pagado": val(idx_sub_pag),
            "monto_cotizacion_previsional_pagado": val(idx_cot_pag),
            "monto_previsional_salud_pagado": val(idx_prev_pag),
            "total_pagado": val(idx_total_pag),
        }
    else:
        # Patrón 2024
        # Primer bloque (Sistema): Monto de Subsidio | Monto cotizacion previsional | Monto Previsional Salud | Total
        # Segundo bloque (Pagado): idem, segunda ocurrencia
        idx_sub_sist = find_first("monto de subsidio")
        idx_cot_sist = find_first("cotizacion previsional")
        idx_prev_sist = find_first("monto previsional salud")
        idx_total_sist = find_first("total")

        idx_sub_pag = find_second("monto de subsidio")
        idx_cot_pag = find_second("cotizacion previsional")
        idx_prev_pag = find_second("monto previsional salud")
        idx_total_pag = find_second("total")

        return {
            "monto_subsidio_sistema": val(idx_sub_sist),
            "monto_cotizacion_previsional_sistema": val(idx_cot_sist),
            "monto_previsional_salud_sistema": val(idx_prev_sist),
            "total_sistema": val(idx_total_sist),
            "monto_subsidio_pagado": val(idx_sub_pag),
            "monto_cotizacion_previsional_pagado": val(idx_cot_pag),
            "monto_previsional_salud_pagado": val(idx_prev_pag),
            "total_pagado": val(idx_total_pag),
        }


# LECTORES

def leer_fuente(data_bytes: bytes):
    """Lee la planilla madre de licencias y devuelve sus tres insumos.

    Estructura esperada del libro (RB-14, RB-16):

    * Hoja ``DATOS``: un funcionario por fila; encabezados en la fila 1
      (``RUN``, ``Nombre``, ``Centro de Costo``, etc.). Es la fuente de
      ``Dim_Funcionario``.
    * Hoja ``LM01-2024``: se reutiliza solo para capturar la columna
      ``Unidad`` (encabezados en fila 2) como fuente adicional de
      establecimientos crudos.
    * Hojas ``LM*`` (excepto ``DATOS``): tablas de hechos de licencias. La
      fila de encabezado se detecta entre las filas 1 y 2 buscando una
      celda que contenga "rut" (RB-14).

    Se abre con ``data_only=True``: se leen los valores calculados por
    Excel, no las fórmulas (RB-16). Las fórmulas rotas llegan como errores
    cacheados y se tratan como vacío aguas arriba (RB-03).

    Args:
        data_bytes: Contenido binario del ``.xlsx`` de la planilla madre.

    Returns:
        Tupla ``(funcionarios, establecimientos_raw, hojas_hechos)``:

        * ``funcionarios``: dict ``{rut: {...}}`` con los datos personales y
          el ``establecimiento`` (Centro de Costo) crudo de cada uno.
        * ``establecimientos_raw``: lista ordenada de nombres crudos de
          establecimiento/unidad observados.
        * ``hojas_hechos``: lista de tuplas ``(nombre_hoja, worksheet,
          fila_header, headers)`` para cada hoja de hechos.

    Raises:
        KeyError: Si el libro no contiene las hojas obligatorias ``DATOS``
            o ``LM01-2024``.
    """
    wb = openpyxl.load_workbook(io.BytesIO(data_bytes), data_only=True)
    datos = wb["DATOS"]
    h = [c.value for c in datos[1]]

    # Precalcular índices de columnas (una sola vez)
    c = {name: i for i, name in enumerate(h)}

    funcionarios = {}
    establecimientos_raw = set()
    for row in datos.iter_rows(min_row=2, values_only=True):
        rut = norm_rut(row[c.get("RUN")])
        if not rut:
            # RB-15: filas sin RUT válido no constituyen funcionario.
            continue
        cc = (row[c.get("Centro de Costo")] or "").strip() or None
        if cc:
            establecimientos_raw.add(cc)
        funcionarios[rut] = {
            "rut": rut,
            "nombre": (row[c.get("Nombre")] or "").strip(),
            "fecha_nacimiento": row[c.get("Fecha Nacimiento")],
            "sexo": row[c.get("Sexo")],
            "estado_civil": row[c.get("Estado Civil")],
            "direccion": row[c.get("Dirección")],
            "comuna": row[c.get("Comuna")],
            "telefono": row[c.get("Teléfono")],
            "telefono_emergencia": row[c.get("Teléfono Emergencia")],
            "nacionalidad": row[c.get("Nacionalidad")],
            "formacion_profesional": row[c.get("Formación Profesional")],
            "cargo": row[c.get("Cargo")],
            "establecimiento": cc,
        }

    lm1 = wb["LM01-2024"]
    h1 = [c.value for c in lm1[2]]
    if "Unidad" in h1:
        idxu = h1.index("Unidad")
        for row in lm1.iter_rows(min_row=3, values_only=True):
            v = row[idxu]
            if v:
                establecimientos_raw.add(str(v).strip())

    hojas_hechos = []
    for name in wb.sheetnames:
        if not name.startswith("LM") or name == "DATOS":
            continue
        ws = wb[name]
        for hr in (1, 2):
            headers = [c.value for c in ws[hr]]
            if any(v and "rut" in str(v).lower() for v in headers):
                hojas_hechos.append((name, ws, hr, headers))
                break

    return funcionarios, sorted(establecimientos_raw), hojas_hechos


def leer_dim_establecimiento(data_bytes: bytes) -> list:
    """Lee la tabla maestra ``Establecimientos.xlsx``.

    La fila de encabezado se detecta dentro de las primeras 4 filas
    buscando la celda "Tipo". Se esperan (todas opcionales salvo el
    nombre) las columnas: ``Tipo``, ``Nombre establecimiento``, ``Comuna``,
    ``Dirección``, ``Telefono`` y ``Sitio web``.

    Los nombres duplicados se descartan conservando la primera aparición.

    Args:
        data_bytes: Contenido binario del ``.xlsx`` maestro.

    Returns:
        Lista de dicts ``{"tipo", "establecimiento", "comuna", "direccion",
        "telefono", "sitio_web"}``, sin duplicados por nombre. Lista vacía
        si no se encontró la fila de encabezado.
    """
    wb = openpyxl.load_workbook(io.BytesIO(data_bytes), data_only=True)
    ws = wb.active
    header_row = 1
    headers = []
    for r in range(1, 5):
        row_vals = [c.value for c in ws[r]]
        if any(v and "tipo" in str(v).lower() for v in row_vals):
            headers = row_vals
            header_row = r
            break

    idx_tipo = headers.index("Tipo") if "Tipo" in headers else None
    idx_nombre = headers.index("Nombre establecimiento") if "Nombre establecimiento" in headers else None
    idx_comuna = headers.index("Comuna") if "Comuna" in headers else None
    idx_direccion = headers.index("Dirección") if "Dirección" in headers else None
    idx_telefono = headers.index("Telefono") if "Telefono" in headers else None
    idx_web = headers.index("Sitio web") if "Sitio web" in headers else None

    establecimientos = []
    vistos = set()
    for r in range(header_row + 1, ws.max_row + 1):
        tipo = ws.cell(r, idx_tipo + 1).value if idx_tipo is not None else None
        nombre = ws.cell(r, idx_nombre + 1).value if idx_nombre is not None else None
        if not nombre or not str(nombre).strip():
            continue
        nombre_limpio = str(nombre).strip()
        if nombre_limpio in vistos:
            continue
        vistos.add(nombre_limpio)
        establecimientos.append({
            "tipo": tipo,
            "establecimiento": nombre_limpio,
            "comuna": ws.cell(r, idx_comuna + 1).value if idx_comuna is not None else None,
            "direccion": ws.cell(r, idx_direccion + 1).value if idx_direccion is not None else None,
            "telefono": ws.cell(r, idx_telefono + 1).value if idx_telefono is not None else None,
            "sitio_web": ws.cell(r, idx_web + 1).value if idx_web is not None else None,
        })
    return establecimientos


# TRANSFORMACIÓN

def construir_dim_afp(hojas_hechos: list) -> list:
    """Construye la dimensión de AFPs con sus tasas a partir del histórico.

    Reglas de negocio (RB-07):

    * Solo entran las AFPs canónicas válidas (valores de ``AFP_MAP`` más el
      literal ``"No Aplica"``); los valores no reconocidos se ignoran aquí
      (ya fueron marcados en la migración de hechos).
    * Si una AFP aparece a veces con tasa explícita y a veces sin ella, las
      filas sin tasa heredan la **tasa conocida** de esa AFP observada en
      cualquier hoja.
    * Si jamás se observó tasa para una AFP, se registra con tasa ``-1``
      como señal de "dato faltante a completar".
    * Se fuerzan dos registros fijos: ``("No Aplica", 0)`` y
      ``("Pensionado (no aplica AFP)", 0)``, para que las validaciones de la
      planilla de hechos siempre tengan esas opciones disponibles.

    Args:
        hojas_hechos: Lista de tuplas ``(nombre, worksheet, fila_header,
            headers)`` de :func:`leer_fuente`.

    Returns:
        Lista de dicts ``{"afp": str, "tasa": float}``, una por combinación
        AFP+tasa observada, ordenada por AFP y tasa descendente.
    """
    # AFPs canónicas válidas: las del mapa + No Aplica
    afps_validas = set(AFP_MAP.values()) | {"No Aplica"}

    # Paso 1: extraer tasas conocidas por AFP (de cualquier hoja)
    tasas_conocidas = {}
    for _, ws2, hrow2, headers2 in hojas_hechos:
        try:
            idx2 = headers2.index("A.F.P.")
        except ValueError:
            continue
        for row in ws2.iter_rows(min_row=hrow2 + 1, values_only=True):
            canon2, tasa2, _ = clasificar_afp(row[idx2])
            if canon2 and canon2 in afps_validas and "no aplica" not in canon2.lower() and tasa2 is not None:
                tasas_conocidas[canon2] = tasa2

    # Paso 2: indexar solo AFPs válidas observadas
    vistos = {}
    for _, ws, hrow, headers in hojas_hechos:
        try:
            idx = headers.index("A.F.P.")
        except ValueError:
            continue
        for row in ws.iter_rows(min_row=hrow + 1, values_only=True):
            canon, tasa, _ = clasificar_afp(row[idx])
            if canon and canon in afps_validas and "no aplica" not in canon.lower():
                if tasa is None:
                    tasa = tasas_conocidas.get(canon, -1)
                vistos[(canon, tasa)] = vistos.get((canon, tasa), 0) + 1

    # Asegurar registros fijos en Dim_AFP
    vistos[("No Aplica", 0)] = vistos.get(("No Aplica", 0), 0)
    vistos[("Pensionado (no aplica AFP)", 0)] = vistos.get(("Pensionado (no aplica AFP)", 0), 0)

    return [{"afp": k[0], "tasa": k[1]} for k, _ in sorted(vistos.items(), key=lambda x: (x[0][0], -x[0][1]))]


def migrar_hechos(hojas_hechos: list, funcionarios: dict, catalogo_norm: dict, catalogo_patterns: dict,
                   TIPO_LICENCIA_CANON: list, INSTITUCION_SALUD_CANON: list, RESOLUCION_MEDICA_CANON: list):
    """Migra y consolida las filas de todas las hojas LM en hechos únicos.

    Es el corazón del pipeline. Por cada fila de cada hoja de hechos:

    1. **Identificación del funcionario** (RB-09): primero por RUT
       normalizado; si no aparece, por coincidencia difusa de nombre
       (cutoff 0.85); si aún no, se crea un *placeholder* en
       ``Dim_Funcionario`` para no dejar RUTs huérfanos y se marca la
       inconsistencia.
    2. **Canonización de dimensiones** (RB-04 a RB-08): establecimiento
       (con *fallback* al Centro de Costo del funcionario cuando la fila no
       lo trae), tipo de licencia, institución de salud, AFP+tasa
       (RB-06/RB-07) y resolución médica (con *fallback* a la columna
       "Estado" si la de resolución está vacía).
    3. **Montos dobles** (RB-11): :func:`extraer_montos_dobles`.
    4. **Validaciones**: fecha de término anterior a la de inicio se marca
       como inconsistencia, pero la fila se conserva (RB-13).
    5. **Deduplicación entre hojas** (RB-10): la clave es
       ``(RUT, Folio, Fecha Inicio)``; si el mismo evento aparece en varias
       hojas, gana la fuente de **mayor año** (extraído del nombre de la
       hoja, ej. ``LM03-2025`` -> 2025).

    Toda inconsistencia detectada se acumula en el campo
    ``estado_migracion`` (``"OK"`` si no hubo ninguna), que termina en la
    columna "Detalle inconsistencia" del archivo de salida.

    Filas sin RUT **ni** folio se descartan (RB-15).

    Args:
        hojas_hechos: Hojas de hechos de :func:`leer_fuente`.
        funcionarios: Dict ``{rut: {...}}`` de :func:`leer_fuente`. Se
            **muta**: los placeholders de funcionarios no encontrados se
            agregan aquí para que aparezcan en ``Dim_Funcionario``.
        catalogo_norm: ``{nombre_normalizado: nombre_canonico}`` del
            maestro de establecimientos.
        catalogo_patterns: Regex de establecimientos (``RE_ESTABLECIMIENTO``).
        TIPO_LICENCIA_CANON: Lista canónica de tipos de licencia.
        INSTITUCION_SALUD_CANON: Lista canónica de instituciones.
        RESOLUCION_MEDICA_CANON: Lista canónica de estados de resolución.

    Returns:
        Tupla ``(hechos, nuevos_establecimientos)``:

        * ``hechos``: lista de dicts de hechos deduplicados, con los campos
          de la tabla ``Hechos_Licencias`` (incluye los 8 montos, ``origen``
          con el nombre de la hoja fuente y ``estado_migracion``).
        * ``nuevos_establecimientos``: dict ``{nombre_crudo: True}`` con los
          establecimientos no presentes en el maestro (se agregan a
          ``Dim_Establecimiento`` como tipo "Otro").
    """
    nuevos_establecimientos = {}
    salida = {}     # dedup_key -> registro (quedarse con la fuente más reciente)
    vistos = {}     # dedup_key -> año de la fuente
    tasas_conocidas = {} # Precalcular tasas conocidas por AFP para filas sin tasa explícita

    for _, ws2, hrow2, headers2 in hojas_hechos:
        try:
            idx2 = headers2.index("A.F.P.")
        except ValueError:
            continue
        for row in ws2.iter_rows(min_row=hrow2 + 1, values_only=True):
            canon2, tasa2, _ = clasificar_afp(row[idx2])
            if canon2 and "no aplica" not in canon2.lower() and tasa2 is not None:
                tasas_conocidas[canon2] = tasa2

    for name, ws, hrow, headers in hojas_hechos:
        # --- PRECÁLCULO DE ÍNDICES (una sola vez por hoja) ---
        def _idx(*names):
            for n in names:
                if n in headers:
                    return headers.index(n)
            return None

        ix = {
            "rut": _idx("Rut"),
            "folio": _idx("Folio licencia", "Folio Minsal"),
            "nombre": _idx("Nombre"),
            "fecha_ini": _idx("Fecha Inicio", "Fech. Inicio"),
            "fecha_ter": _idx("Fecha Termino", "Fech. Termino"),
            "dias_lm": _idx("Días LM", "Días Lic"),
            "tipo": _idx("Tipo Licencia"),
            "inst": _idx("Institución Salud", "Institucion Salud"),
            "afp": _idx("A.F.P."),
            "resol": _idx("Resolución Médica", "Resolucion Medica"),
            "estado": _idx("Estado"),
            "obs1": _idx("Observaciones"),
            "obs2": _idx("Observaciones 2"),
            "sexo": _idx("Sexo"),
            "estab": _idx("Estableciemiento", "Establecimiento", "Unidad",
                           "Centro de Costo", "Lugar", "Sede", "Ubicacion"),
        }

        for row in ws.iter_rows(min_row=hrow + 1, values_only=True):
            rut_raw = row[ix["rut"]] if ix["rut"] is not None else None
            if rut_raw is None and (ix["folio"] is None or row[ix["folio"]] is None):
                continue

            rut = norm_rut(rut_raw)
            nombre_raw = row[ix["nombre"]] if ix["nombre"] is not None else None
            estado_migracion = []
            func = funcionarios.get(rut) if rut else None

            if func is None and nombre_raw:
                nombres_idx = {norm(f["nombre"]): k for k, f in funcionarios.items()}
                cand = get_close_matches(norm(nombre_raw), nombres_idx.keys(), n=1, cutoff=0.85)
                if cand:
                    rut = nombres_idx[cand[0]]
                    func = funcionarios.get(rut)

            # Crear placeholder en Dim_Funcionario para evitar RUTs huérfanos
            if func is None:
                estado_migracion.append("RUT/Nombre no encontrado en Dim_Funcionario")
                if rut and rut not in funcionarios:
                    funcionarios[rut] = {
                        "rut": rut,
                        "nombre": (nombre_raw or "").strip(),
                        "fecha_nacimiento": None,
                        "sexo": None,
                        "estado_civil": None,
                        "direccion": None,
                        "comuna": None,
                        "telefono": None,
                        "telefono_emergencia": None,
                        "nacionalidad": None,
                        "formacion_profesional": None,
                        "cargo": None,
                        "establecimiento": None,
                    }
                func = funcionarios.get(rut)

            # Establecimiento: buscar en múltiples columnas posibles
            estab_raw = row[ix["estab"]] if ix["estab"] is not None else None
            estab_canon = None
            if estab_raw:
                estab_canon, est_estado = clasificar_establecimiento(estab_raw, catalogo_norm, catalogo_patterns)
                if est_estado.startswith("NUEVO"):
                    nuevos_establecimientos[estab_canon] = True
                    estado_migracion.append(est_estado)
                elif est_estado != "OK" and not est_estado.startswith("OK"):
                    estado_migracion.append("Establecimiento: " + est_estado)
            else:
                # Fallback: usar el del funcionario; si es None, dejarlo como None
                estab_canon = func["establecimiento"] if func else None
                if estab_canon:
                    estab_canon2, est_estado2 = clasificar_establecimiento(estab_canon, catalogo_norm, catalogo_patterns)
                    if est_estado2.startswith("NUEVO"):
                        nuevos_establecimientos[estab_canon2] = True
                    estab_canon = estab_canon2

            tipo_canon, tipo_estado = clasificar_generico(
                row[ix["tipo"]] if ix["tipo"] is not None else None, TIPO_LICENCIA_CANON, RE_TIPO)
            if tipo_estado not in ("OK", "Vacio"):
                estado_migracion.append("Tipo Licencia: " + tipo_estado)

            inst_canon, inst_estado = clasificar_generico(
                row[ix["inst"]] if ix["inst"] is not None else None, INSTITUCION_SALUD_CANON, RE_INST)
            if inst_estado not in ("OK", "Vacio"):
                estado_migracion.append("Institucion Salud: " + inst_estado)

            afp_canon, tasa, afp_estado = clasificar_afp(
                row[ix["afp"]] if ix["afp"] is not None else None)
            if afp_estado not in ("OK", "Vacio"):
                estado_migracion.append("AFP: " + afp_estado)
                afp_canon = "No Aplica"
                tasa = 0
            if tasa is None and afp_canon:
                if "no aplica" in afp_canon.lower():
                    tasa = 0
                else:
                    tasa = tasas_conocidas.get(afp_canon, -1)

            resol_raw = row[ix["resol"]] if ix["resol"] is not None else None
            resol_canon, resol_estado = clasificar_resolucion(resol_raw, RESOLUCION_MEDICA_CANON)
            if resol_canon is None and ix["estado"] is not None:
                resol_canon, resol_estado = clasificar_resolucion(
                    row[ix["estado"]], RESOLUCION_MEDICA_CANON)
            if resol_estado not in ("OK", "Vacio") and not str(resol_estado).startswith("LEGACY"):
                estado_migracion.append("Resolucion Medica: " + resol_estado)

            # Extraer ambos bloques de montos
            montos = extraer_montos_dobles(headers, row)

            # Deduplicación: mismo evento de licencia en fuentes distintas
            # Si un folio aparece en varias hojas, quedarse con la fuente de mayor año.
            # La clave usa RUT + Folio + Fecha Inicio (no incluye Fecha Término ni Fuente).
            folio = row[ix["folio"]] if ix["folio"] is not None else None
            fecha_ini = row[ix["fecha_ini"]] if ix["fecha_ini"] is not None else None
            fecha_ter = row[ix["fecha_ter"]] if ix["fecha_ter"] is not None else None

            dedup_key = (
                str(rut or ""),
                str(folio or ""),
                str(fecha_ini or ""),
            )
            m_anio = re.search(r'(\d{4})', name)
            anio_fuente = int(m_anio.group(1)) if m_anio else 0

            if dedup_key in vistos and anio_fuente <= vistos[dedup_key]:
                continue  # ya existe registro de fuente igual o más reciente
            vistos[dedup_key] = anio_fuente

            if (fecha_ini and fecha_ter and isinstance(fecha_ini, datetime)
                    and isinstance(fecha_ter, datetime) and fecha_ter < fecha_ini):
                estado_migracion.append("Fecha Termino anterior a Fecha Inicio")

            salida[dedup_key] = {
                "rut": rut or (norm_rut(rut_raw) or ""),
                "nombre": func["nombre"] if func else (nombre_raw or ""),
                "fecha_nacimiento": func["fecha_nacimiento"] if func else None,
                "sexo": func["sexo"] if func else (row[ix["sexo"]] if ix["sexo"] is not None else None),
                "establecimiento": estab_canon,
                "folio_licencia": folio,
                "fecha_inicio": fecha_ini,
                "fecha_termino": fecha_ter,
                "dias_lm": row[ix["dias_lm"]] if ix["dias_lm"] is not None else None,
                "tipo_licencia": tipo_canon,
                "institucion_salud": inst_canon,
                "afp": afp_canon or "No Aplica",
                "tasa_afp": tasa if tasa is not None else 0,
                "resolucion_medica": resol_canon,
                **montos,
                "observaciones": " | ".join(
                    str(x) for x in [
                        row[ix["obs1"]] if ix["obs1"] is not None else None,
                        row[ix["obs2"]] if ix["obs2"] is not None else None,
                    ] if x
                ),
                "origen": name,
                "estado_migracion": "; ".join(estado_migracion) if estado_migracion else "OK",
            }

    return list(salida.values()), nuevos_establecimientos


def migrar_descuentos(hojas_hechos: list) -> list:
    """Extrae los descuentos por período de todas las hojas LM (RB-12).

    Las hojas de hechos llevan los descuentos en formato **ancho**: una
    columna por mes con encabezado ``MONTO DESCONTADO <MES> <AÑO>`` (ej.
    "MONTO DESCONTADO MARZO 2025"). Esta función los pivotea a formato
    **largo**: una fila por (folio, RUT, período) con su monto.

    Reglas de negocio:

    * El período se normaliza a ``YYYY-MM`` usando ``MESES_MAP``.
    * Solo se emiten montos **distintos de cero** (las celdas vacías o en 0
      no generan registro).
    * Filas sin RUT ni folio se omiten (RB-15).
    * Valores no numéricos se ignoran silenciosamente.

    Args:
        hojas_hechos: Hojas de hechos de :func:`leer_fuente`.

    Returns:
        Lista de dicts ``{"folio_licencia", "rut", "periodo",
        "monto_descuento", "origen"}`` para la tabla ``Hechos_Descuentos``.
    """
    descuentos = []
    re_desc = re.compile(r"MONTO\s+DESCONTADO\s+(\w+)\s+(\d{4})", re.IGNORECASE)

    for name, ws, hrow, headers in hojas_hechos:
        # 1) Detectar qué columnas son de descuento y a qué período corresponden
        desc_cols = []   # [(índice_en_row, periodo YYYY-MM), ...]
        for idx, h in enumerate(headers):
            if h and isinstance(h, str):
                m = re_desc.match(h.strip())
                if m:
                    mes_nombre = m.group(1).upper()
                    anio = m.group(2)
                    mes_num = MESES_MAP.get(mes_nombre)
                    if mes_num:
                        desc_cols.append((idx, f"{anio}-{mes_num}"))

        if not desc_cols:
            continue

        # 2) Precalcular índices de columnas fijas (evita get_by_any en cada fila)
        def _find_col(*names):
            for n in names:
                try:
                    return headers.index(n)
                except ValueError:
                    pass
            return None

        idx_rut   = _find_col("Rut")
        idx_folio = _find_col("Folio licencia", "Folio Minsal")

        # 3) Recorrer filas con iter_rows (mucho más rápido en WASM/Pyodide)
        for row in ws.iter_rows(min_row=hrow + 1, values_only=True):
            rut_raw = row[idx_rut] if idx_rut is not None else None
            folio   = row[idx_folio] if idx_folio is not None else None
            rut = norm_rut(rut_raw) if rut_raw else ""

            # Si no hay ni RUT ni Folio, saltamos la fila
            if not rut and not folio:
                continue

            for idx, periodo in desc_cols:
                valor = row[idx]
                if valor is None:
                    continue
                try:
                    monto = float(valor)
                    if monto != 0:
                        descuentos.append({
                            "folio_licencia": folio,
                            "rut": rut,
                            "periodo": periodo,
                            "monto_descuento": monto,
                            "origen": name,
                        })
                except (ValueError, TypeError):
                    pass

    return descuentos


# WRITERS

# Estilos compartidos por todas las salidas: fuente base 9pt, encabezado
# azul corporativo con texto blanco, bordes finos grises y congelación de la
# fila de títulos (ver _estilizar_header).
FONT = Font(name="Calibri", size=9)
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _estilizar_header(ws, ncols: int) -> None:
    """Aplica el estilo corporativo a la fila de encabezado y la congela.

    Fuente blanca en negrita sobre azul ``1F4E78``, texto centrado con
    ajuste, bordes, alto de fila 30 y paneles congelados en la fila 2.
    """
    for c in range(1, ncols + 1):
        cell = ws.cell(1, c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.freeze_panes = ws.cell(2, 1).coordinate
    ws.row_dimensions[1].height = 30


def _autoancho(ws, max_w: int = 30) -> None:
    """Ajusta el ancho de cada columna a su contenido visible estimado.

    Calcula el largo "renderizado" de cada celda considerando su formato
    numérico (fechas como ``DD-MM-YYYY``, miles con separador, porcentajes,
    signos de moneda y negativos entre paréntesis) y fija el ancho al máximo
    encontrado + 2, acotado a ``max_w``. Las fórmulas (cadenas que empiezan
    con ``=``) no se miden.

    Args:
        ws: Hoja de cálculo a ajustar.
        max_w: Ancho máximo permitido por columna (30 por defecto).
    """
    def _clean(s):
        if not s:
            return ""
        s = re.sub(r"[\r\n\t]+", " ", str(s))
        s = re.sub(r"\s+", " ", s)
        return s.strip()

    def _vlen(cell):
        v = cell.value
        if v is None:
            return 0
        if isinstance(v, str) and v.startswith("="):
            return 0
        f = cell.number_format or "General"
        if isinstance(v, (datetime, date)):
            if "YYYY-MM-DD" in f:
                return 10
            if "DD-MM-YYYY" in f:
                return 10
            if "DD/MM/YYYY" in f:
                return 10
            return len(v.strftime("%d-%m-%Y"))
        if isinstance(v, (int, float)):
            m = re.search(r"\.(0+)", f)
            d = len(m.group(1)) if m else 0
            p, t = "%" in f, "#,##0" in f or "#,###" in f
            n = v * 100 if p else v
            s = (f"{n:,.{d}f}" if t else f"{n:.{d}f}") + ("%" if p else "")
            if v < 0 and ";" in f:
                s = f"({s.lstrip('-')})"
            return len(s) + (1 if any(c in f for c in "$€£¥") else 0)
        return len(_clean(v))

    real_max_row = 1
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        if any(cell.value is not None for cell in row):
            real_max_row = row[0].row
    ml = {}
    for col in ws.iter_cols(min_row=1, max_row=real_max_row):
        for cell in col:
            c = cell.column
            l = _vlen(cell)
            if l > ml.get(c, 0):
                ml[c] = l
    for c, l in ml.items():
        ws.column_dimensions[get_column_letter(c)].width = min(l + 2, max_w)


def _escribir_dim(title: str, headers: list, rows: list, num_fmt: dict = None) -> bytes:
    """Escribe un archivo de dimensión completo (libro de una sola hoja).

    Genera la hoja con encabezado estilizado, bordes y fuente base en todas
    las celdas, formatos numéricos por columna (``num_fmt``: dict
    ``{nro_columna: formato}``, 1-indexado), autoancho y una **tabla de
    Excel** nativa (filtros y referencias estructuradas para Power BI).

    Args:
        title: Nombre de la hoja (se reutiliza, sin espacios, como nombre de
            la tabla de Excel).
        headers: Encabezados de columna.
        rows: Filas de datos.
        num_fmt: Formatos numéricos opcionales por columna.

    Returns:
        Contenido binario del ``.xlsx`` generado.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = title
    ws.append(headers)
    for r in rows:
        ws.append(r)
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(r, c)
            cell.font = FONT
            cell.border = BORDER
            if num_fmt and c in num_fmt:
                cell.number_format = num_fmt[c]
    _estilizar_header(ws, len(headers))
    ws.row_dimensions[1].height = None
    _autoancho(ws)
    ws.add_table(Table(displayName=title.replace(" ", ""), ref=f"A1:{get_column_letter(len(headers))}{ws.max_row}"))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _add_ref(wb: Workbook, name: str, headers: list, rows: list) -> int:
    """Agrega una hoja de referencia **oculta** para las fórmulas.

    Las hojas ``ref_*`` son copias internas de las dimensiones que alimentan
    las fórmulas ``INDEX/MATCH`` y las listas de validación de
    ``Hechos_Licencias``; no se editan a mano, se regeneran.

    Returns:
        Número de filas escritas (incluido el encabezado), usado para acotar
        los rangos de las fórmulas.
    """
    ws = wb.create_sheet(name)
    ws.append(headers)
    for r in rows:
        ws.append(r)
    ws.sheet_state = "hidden"
    return ws.max_row


def _add_dv(ws, col: int, formula: str, n_rows: int, title: str = "Valor no valido",
            msg: str = "Debe seleccionar un elemento de la lista.") -> None:
    """Agrega una validación de datos tipo lista desplegable a una columna.

    Args:
        ws: Hoja destino.
        col: Índice de columna (1-indexado).
        formula: Fórmula del rango de la lista (ej. ``"=ref_Listas!$A$2:$A$10"``).
        n_rows: Cantidad de filas a validar desde la fila 2.
        title: Título del cuadro de error.
        msg: Mensaje de error cuando el valor no pertenece a la lista.
    """
    dv = DataValidation(type="list", formula1=formula, allow_blank=True, showErrorMessage=True, showInputMessage=True)
    dv.error, dv.errorTitle = msg, title
    dv.prompt, dv.promptTitle = "Seleccione un valor de la lista desplegable", "Lista de valores"
    rng = f"{get_column_letter(col)}2:{get_column_letter(col)}{1 + n_rows}"
    dv.add(rng)
    ws.add_data_validation(dv)


def escribir_dim_funcionario(funcionarios: dict) -> bytes:
    """Genera ``01_Dim_Funcionario.xlsx``, ordenado por nombre.

    Columnas: RUT, Nombre, Fecha Nacimiento (formato ``DD-MM-YYYY``), Sexo,
    Estado Civil, Direccion, Comuna, Telefono, Telefono Emergencia,
    Nacionalidad, Formacion Profesional, Cargo y Establecimiento.
    """
    rows = [
        [f["rut"], f["nombre"], f["fecha_nacimiento"], f["sexo"], f["estado_civil"],
         f["direccion"], f["comuna"], f["telefono"], f["telefono_emergencia"],
         f["nacionalidad"], f["formacion_profesional"], f["cargo"], f["establecimiento"]]
        for f in sorted(funcionarios.values(), key=lambda x: x["nombre"])
    ]
    return _escribir_dim(
        "Funcionario",
        ["RUT", "Nombre", "Fecha Nacimiento", "Sexo", "Estado Civil", "Direccion", "Comuna",
         "Telefono", "Telefono Emergencia", "Nacionalidad", "Formacion Profesional", "Cargo", "Establecimiento"],
        rows, num_fmt={3: "DD-MM-YYYY"},
    )


def escribir_dim_establecimiento(establecimientos: list) -> bytes:
    """Genera ``02_Dim_Establecimiento.xlsx``, ordenado y deduplicado por nombre.

    Columnas: Establecimiento, Tipo, Comuna, Direccion, Telefono, Sitio Web.
    """
    # Deduplicar por nombre
    vistos = set()
    unicos = []
    for e in sorted(establecimientos, key=lambda x: x["establecimiento"]):
        if e["establecimiento"] not in vistos:
            vistos.add(e["establecimiento"])
            unicos.append(e)
    rows = [
        [e["establecimiento"], e["tipo"], e.get("comuna"), e.get("direccion"), e.get("telefono"), e.get("sitio_web")]
        for e in unicos
    ]
    return _escribir_dim("Establecimiento", ["Establecimiento", "Tipo", "Comuna", "Direccion", "Telefono", "Sitio Web"], rows)


def escribir_dim_afp(afp_filas: list) -> bytes:
    """Genera ``03_Dim_AFP.xlsx`` (columnas AFP y Tasa, formato ``0.00``)."""
    rows = [[f["afp"], f["tasa"]] for f in afp_filas]
    return _escribir_dim("AFP", ["AFP", "Tasa"], rows, num_fmt={2: "0.00"})


def escribir_hechos_descuentos(descuentos: list) -> bytes:
    """Genera ``05_Hechos_Descuentos.xlsx`` (tabla de hechos en formato largo).

    Columnas: Folio Licencia, RUT, Periodo (``YYYY-MM``, RB-12), Monto
    Descuento (formato ``#,##0``) y Fuente (hoja de origen). Si no hay
    descuentos, el archivo se emite igual, solo con el encabezado.
    """
    wb = Workbook()
    if descuentos:
        wb.remove(wb.active)
        ws = wb.create_sheet("Hechos_Descuentos")
        headers = ["Folio Licencia", "RUT", "Periodo", "Monto Descuento", "Fuente"]
        ws.append(headers)
        ncols = len(headers)
        C = {h: i + 1 for i, h in enumerate(headers)}

        for d in descuentos:
            ws.append([
                d["folio_licencia"],
                d["rut"],
                d["periodo"],
                d["monto_descuento"],
                d["origen"],
            ])

        _estilizar_header(ws, ncols)
        ws.row_dimensions[1].height = None  
        for r in range(2, ws.max_row + 1):
            for c in range(1, ncols + 1):
                cell = ws.cell(r, c)
                cell.font = FONT
                cell.border = BORDER
            ws.cell(r, C["Monto Descuento"]).number_format = "#,##0"

        _autoancho(ws)
        ws.add_table(Table(
            displayName="HechosDescuentos",
            ref=f"A1:{get_column_letter(ncols)}{ws.max_row}"
        ))
    else:
        wb.active.title = "Hechos_Descuentos"
        wb.active.append(["Folio Licencia", "RUT", "Periodo", "Monto Descuento", "Fuente"])
        _estilizar_header(wb.active, 5)
        wb.active.row_dimensions[1].height = None

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def escribir_hechos(funcionarios: dict, afp_filas: list, hechos: list, TIPO_LICENCIA_CANON: list,
                    INSTITUCION_SALUD_CANON: list, RESOLUCION_MEDICA_CANON: list, descuentos: list = None) -> bytes:
    """Genera ``04_Hechos_Licencias.xlsx``: la planilla madre nueva.

    Libro con tres tipos de hojas:

    * ``Instrucciones``: guía de uso para quien imputa datos.
    * ``Hechos_Licencias`` (visible): tabla principal con los hechos
      migrados **más 40 filas en blanco** para nuevas licencias.
    * ``ref_Funcionario``, ``ref_AFP``, ``ref_Listas`` y ``ref_Descuentos``
      (ocultas): copias internas que alimentan fórmulas y validaciones.

    Mecánicas de autorrelleno para las filas nuevas:

    * ``Nombre``, ``Fecha Nacimiento`` y ``Sexo`` se calculan con
      ``INDEX/MATCH`` desde ``ref_Funcionario`` a partir del RUT. El
      ``Establecimiento``, en cambio, se escribe como valor directo del
      hecho migrado (decisión de diseño: el centro de costo del funcionario
      puede cambiar en el tiempo y el hecho debe conservar el suyo).
    * ``Tasa AFP``: valor histórico directo en filas migradas; fórmula
      ``INDEX/MATCH`` sobre ``ref_AFP`` en filas nuevas.
    * ``Aplica Descuentos``: fórmula con ``COUNTIF`` + ``HYPERLINK`` que
      enlaza al archivo ``05_Hechos_Descuentos.xlsx`` cuando el folio tiene
      descuentos asociados.
    * Validaciones: listas desplegables para Tipo Licencia, Institución
      Salud, Resolución Médica y A.F.P. (sobre ``ref_Listas``), validación
      de RUT contra ``ref_Funcionario`` y validación de fechas entre
      01-01-2020 y 31-12-2100 (seriales 43831 y 73415).

    El encabezado usa un código de colores por bloque semántico: verde
    (funcionario), azul (licencia), rojo (montos Sistema), fucsia (montos
    Pagado) y naranjo (observaciones/trazabilidad).

    Args:
        funcionarios: Dict ``{rut: {...}}`` (incluye placeholders).
        afp_filas: Filas de ``Dim_AFP`` (``{"afp", "tasa"}``).
        hechos: Hechos migrados de :func:`migrar_hechos`.
        TIPO_LICENCIA_CANON, INSTITUCION_SALUD_CANON, RESOLUCION_MEDICA_CANON:
            Listas canónicas para las validaciones desplegables.
        descuentos: Descuentos de :func:`migrar_descuentos` (opcional).

    Returns:
        Contenido binario del ``.xlsx`` generado.
    """
    wb = Workbook()
    hoja = wb.active
    hoja.title = "Hechos_Licencias"

    # refs ocultas (igual que antes)
    func_rows = [
        [f["rut"], f["nombre"], f["fecha_nacimiento"], f["sexo"], f["establecimiento"]]
        for f in sorted(funcionarios.values(), key=lambda x: x["nombre"])
    ]
    n_func = _add_ref(wb, "ref_Funcionario", ["RUT", "Nombre", "Fecha Nacimiento", "Sexo", "Establecimiento"], func_rows)

    afp_rows = [[f["afp"], f["tasa"]] for f in afp_filas]
    n_afp = _add_ref(wb, "ref_AFP", ["AFP", "Tasa"], afp_rows)

    ref_listas = wb.create_sheet("ref_Listas")
    ref_listas.append(["Tipo Licencia", "Institucion Salud", "Resolucion Medica", "AFP"])
    for i, v in enumerate(TIPO_LICENCIA_CANON, 2):
        ref_listas.cell(i, 1, v)
    for i, v in enumerate(INSTITUCION_SALUD_CANON, 2):
        ref_listas.cell(i, 2, v)
    for i, v in enumerate(RESOLUCION_MEDICA_CANON, 2):
        ref_listas.cell(i, 3, v)
    for i, v in enumerate(["No Aplica"] + [f["afp"] for f in afp_filas], 2):
        ref_listas.cell(i, 4, v)
    ref_listas.sheet_state = "hidden"

    n_tipos = 1 + len(TIPO_LICENCIA_CANON)
    n_inst = 1 + len(INSTITUCION_SALUD_CANON)
    n_resol = 1 + len(RESOLUCION_MEDICA_CANON)
    n_afp_names = 1 + len(afp_filas) + 1

    # ref_Descuentos (hoja oculta, igual patrón que ref_Funcionario)
    n_desc = 0
    if descuentos:
        desc_rows = [
            [d["folio_licencia"], d["rut"], d["periodo"], d["monto_descuento"], d["origen"]]
            for d in descuentos
        ]
        n_desc = _add_ref(wb, "ref_Descuentos", ["Folio Licencia", "RUT", "Periodo", "Monto Descuento", "Fuente"], desc_rows)

    # Headers de Hechos_Licencias (con 8 columnas de montos: 4 Sistema + 4 Pagado)
    headers = [
        "Folio Licencia", "RUT", "Nombre", "Fecha Nacimiento", "Sexo", "Establecimiento",
        "Fecha Inicio", "Fecha Termino", "Dias LM", "Tipo Licencia", "Institucion Salud",
        "A.F.P.", "Tasa AFP", "Resolucion Medica",
        "Monto Subsidio Sistema", "Monto Cotizacion Previsional Sistema", "Monto Previsional Salud Sistema", "Total Sistema",
        "Monto Subsidio Pagado", "Monto Cotizacion Previsional Pagado", "Monto Previsional Salud Pagado", "Total Pagado",
        "Observaciones", "Aplica Descuentos",
        "Fuente", "Detalle inconsistencia"
    ]
    hoja.append(headers)
    ncols = len(headers)
    C = {h: i + 1 for i, h in enumerate(headers)}
    n_hist, n_rows = len(hechos), len(hechos) + 40
    FECHA_MAP = {"fecha_inicio": "Fecha Inicio", "fecha_termino": "Fecha Termino"}

    # Columnas de montos para aplicar formato numérico
    COLS_MONTO_SISTEMA = [
        C["Monto Subsidio Sistema"], C["Monto Cotizacion Previsional Sistema"],
        C["Monto Previsional Salud Sistema"], C["Total Sistema"]
    ]
    COLS_MONTO_PAGADO = [
        C["Monto Subsidio Pagado"], C["Monto Cotizacion Previsional Pagado"],
        C["Monto Previsional Salud Pagado"], C["Total Pagado"]
    ]

    for i in range(n_rows):
        r = i + 2
        h = hechos[i] if i < n_hist else None
        rut_val = h["rut"] if h else None
        hoja.cell(r, C["RUT"], rut_val)
        f_rut = f"{get_column_letter(C['RUT'])}{r}"

        # Establecimiento se escribe DIRECTAMENTE desde el hecho, no como fórmula
        # Solo Nombre, Fecha Nacimiento y Sexo usan fórmula INDEX/MATCH
        for col_name, ref_col in [("Nombre", "B"), ("Fecha Nacimiento", "C"), ("Sexo", "D")]:
            formula = (
                '=IFERROR(INDEX(ref_Funcionario!$' + ref_col + '$2:$' + ref_col + '$' + str(n_func)
                + ',MATCH(' + f_rut + ',ref_Funcionario!$A$2:$A$' + str(n_func) + ',0)),"")'
            )
            hoja.cell(r, C[col_name], formula)
        hoja.cell(r, C["Fecha Nacimiento"]).number_format = "DD-MM-YYYY"

        # Establecimiento: valor directo del hecho (puede ser None)
        if h:
            hoja.cell(r, C["Establecimiento"], h["establecimiento"])

        if h:
            for campo, header in [
                ("folio_licencia", "Folio Licencia"), ("fecha_inicio", "Fecha Inicio"),
                ("fecha_termino", "Fecha Termino"), ("dias_lm", "Dias LM"),
                ("tipo_licencia", "Tipo Licencia"), ("institucion_salud", "Institucion Salud"),
                ("afp", "A.F.P."), ("resolucion_medica", "Resolucion Medica"),
                ("monto_subsidio_sistema", "Monto Subsidio Sistema"),
                ("monto_cotizacion_previsional_sistema", "Monto Cotizacion Previsional Sistema"),
                ("monto_previsional_salud_sistema", "Monto Previsional Salud Sistema"),
                ("total_sistema", "Total Sistema"),
                ("monto_subsidio_pagado", "Monto Subsidio Pagado"),
                ("monto_cotizacion_previsional_pagado", "Monto Cotizacion Previsional Pagado"),
                ("monto_previsional_salud_pagado", "Monto Previsional Salud Pagado"),
                ("total_pagado", "Total Pagado"),
                ("observaciones", "Observaciones"), ("origen", "Fuente"),
            ]:
                hoja.cell(r, C[header], h[campo])
            hoja.cell(r, C["Detalle inconsistencia"], h["estado_migracion"] if h["estado_migracion"] != "OK" else "")
            for fkey, hkey in FECHA_MAP.items():
                if isinstance(h[fkey], datetime):
                    hoja.cell(r, C[hkey]).number_format = "DD-MM-YYYY"

        # Formato numérico para montos
        for col_monto in COLS_MONTO_SISTEMA + COLS_MONTO_PAGADO:
            hoja.cell(r, col_monto).number_format = "#,##0"

        # fórmula Aplica Descuentos con HYPERLINK externo
        if descuentos and n_desc > 0:
            folio_col = get_column_letter(C["Folio Licencia"])
            folio_cell = f"{folio_col}{r}"
            formula = (
                f'=IF(AND({folio_cell}<>"",COUNTIF(ref_Descuentos!$A:$A,{folio_cell})>0),'
                f'HYPERLINK("[05_Hechos_Descuentos.xlsx]Hechos_Descuentos!A"&MATCH({folio_cell},ref_Descuentos!$A:$A,0),'
                f'"Sí ("&COUNTIF(ref_Descuentos!$A:$A,{folio_cell})&" desc.)"),"No")'
            )
            hoja.cell(r, C["Aplica Descuentos"], formula)
        else:
            hoja.cell(r, C["Aplica Descuentos"], "No")

        # Tasa AFP automática. Para filas históricas escribir el valor directo
        # (dato histórico); para filas nuevas (blancas) dejar la fórmula INDEX/MATCH.
        if i < n_hist and h and h.get("tasa_afp") is not None:
            hoja.cell(r, C["Tasa AFP"], h["tasa_afp"])
        else:
            f_afp = f"{get_column_letter(C['A.F.P.'])}{r}"
            formula_afp = "=IFERROR(INDEX(ref_AFP!$B$2:$B$" + str(n_afp) + ",MATCH(" + f_afp + ",ref_AFP!$A$2:$A$" + str(n_afp) + ",0)),0)"
            hoja.cell(r, C["Tasa AFP"], formula_afp)
        hoja.cell(r, C["Tasa AFP"]).number_format = "0.00"

        for c in range(1, ncols + 1):
            cell = hoja.cell(r, c)
            cell.border = BORDER
            cell.font = FONT

    _estilizar_header(hoja, ncols)
    hoja.row_dimensions[1].height = None
    _autoancho(hoja)
    hoja.add_table(Table(displayName="HechosLicencias", ref=f"A1:{get_column_letter(ncols)}{1 + n_rows}"))

    fills = {
        "2E7D32": ["RUT", "Nombre", "Fecha Nacimiento", "Sexo", "Establecimiento"],
        "1565C0": ["Folio Licencia", "Fecha Inicio", "Fecha Termino", "Dias LM", "Tipo Licencia",
                    "Institucion Salud", "A.F.P.", "Tasa AFP", "Resolucion Medica"],
        "C62828": ["Monto Subsidio Sistema", "Monto Cotizacion Previsional Sistema",
                   "Monto Previsional Salud Sistema", "Total Sistema"],
        "AD1457": ["Monto Subsidio Pagado", "Monto Cotizacion Previsional Pagado",
                   "Monto Previsional Salud Pagado", "Total Pagado"],
        "F57C00": ["Observaciones", "Aplica Descuentos", "Fuente", "Detalle inconsistencia"],
    }
    for color, campos in fills.items():
        fill = PatternFill("solid", start_color=color, end_color=color)
        for campo in campos:
            hoja.cell(1, C[campo]).fill = fill

    # Validaciones de listas (igual que antes)
    _add_dv(hoja, C["Tipo Licencia"], "=ref_Listas!$A$2:$A$" + str(n_tipos), n_rows)
    _add_dv(hoja, C["Institucion Salud"], "=ref_Listas!$B$2:$B$" + str(n_inst), n_rows)
    _add_dv(hoja, C["Resolucion Medica"], "=ref_Listas!$C$2:$C$" + str(n_resol), n_rows)
    _add_dv(hoja, C["A.F.P."], "=ref_Listas!$D$2:$D$" + str(n_afp_names), n_rows)

    for col_fecha in [C["Fecha Inicio"], C["Fecha Termino"]]:
        dv = DataValidation(type="date", operator="between", formula1="43831", formula2="73415",
                             allow_blank=True, showErrorMessage=True, showInputMessage=True)
        dv.error, dv.errorTitle = "Debe ingresar una fecha valida.", "Fecha invalida"
        dv.prompt, dv.promptTitle = "Haga doble clic o presione Ctrl+; para insertar fecha.", "Seleccion de fecha"
        dv.add(f"{get_column_letter(col_fecha)}2:{get_column_letter(col_fecha)}{1 + n_rows}")
        hoja.add_data_validation(dv)

    dv_rut = DataValidation(type="list", formula1="=ref_Funcionario!$A$2:$A$" + str(n_func),
                             allow_blank=True, showErrorMessage=True, showInputMessage=True)
    dv_rut.error = "El RUT no existe en Dim_Funcionario. Agregue el funcionario en el archivo correspondiente."
    dv_rut.errorTitle = "Funcionario no encontrado"
    dv_rut.prompt = "Seleccione o escriba un RUT registrado en Dim_Funcionario"
    dv_rut.promptTitle = "Validacion de RUT"
    dv_rut.add(f"{get_column_letter(C['RUT'])}2:{get_column_letter(C['RUT'])}{1 + n_rows}")
    hoja.add_data_validation(dv_rut)


    # Instrucciones actualizadas
    inst = wb.create_sheet("Instrucciones", 0)
    inst.sheet_view.showGridLines = False
    inst.column_dimensions["A"].width = 110
    textos = [
        "INSTRUCCIONES - Hechos_Licencias (planilla madre nueva, esquema estrella)",
        "",
        "1) Escriba el RUT del funcionario. Nombre, Fecha de Nacimiento, Sexo se completan automaticamente desde Dim_Funcionario. Establecimiento se escribe directamente desde el hecho migrado.",
        "",
        "2) 'Tipo Licencia', 'Institucion Salud', 'A.F.P.' y 'Resolucion Medica' tienen lista desplegable. Para agregar valores nuevos, actualice la dimension correspondiente y regenere.",
        "",
        "3) Al escribir la A.F.P., la 'Tasa AFP' se completa automaticamente desde Dim_AFP. Use 'No Aplica' para pensionados.",
        "",
        "4) Las filas migradas indican su origen en 'Fuente' y posibles inconsistencias en 'Detalle inconsistencia'.",
        "",
        "5) MONTOS DOBLES: las columnas con sufijo 'Sistema' (fondo rojo) corresponden al calculo Dep/Netcore. Las columnas con sufijo 'Pagado' (fondo fucsia) corresponden a lo efectivamente pagado por FONASA/ISAPRE. En hojas 2025/2026 el primer bloque usa columnas AFP/SALUD; el mapeo se realiza automaticamente.",
        "",
        "6) La columna 'Aplica Descuentos' muestra un enlace si el folio tiene descuentos en 05_Hechos_Descuentos.xlsx. Al hacer clic se abre ese archivo y salta a la primera fila del folio. Desde alli puede filtrar la tabla por Folio para ver todos los descuentos.",
        "",
        "7) Las hojas ref_* son copias internas para las formulas. NO se editan aqui; se actualizan desde los archivos 01/02/03/05 y regenerando.",
        "",
        "8) Para Power BI: importe los 5 archivos y arme relaciones: Hechos.RUT -> Funcionario.RUT, Hechos.Establecimiento -> Establecimiento.Establecimiento, Hechos.A.F.P. -> AFP.AFP.",
    ]
    for i, t in enumerate(textos, 1):
        c = inst.cell(i, 1, t)
        c.font = Font(name="Calibri", bold=(i == 1), size=14 if i == 1 else 11)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        inst.row_dimensions[i].height = 34 if i > 1 and t else 14

    wb.active = 1
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# PIPELINE (punto de entrada)

def procesar(src_bytes: bytes, dim_est_bytes: bytes, log_callback=None, pbit_data: bytes = None) -> dict:
    """Ejecuta el pipeline completo y devuelve todos los archivos generados.

    Orquesta las cuatro etapas (lectura -> clasificación -> migración ->
    escritura) y emite por ``log_callback`` un log legible para la interfaz
    web: conteos de funcionarios y establecimientos, folios repetidos entre
    hojas (advertencia temprana de la deduplicación RB-10), el detalle de la
    clasificación canónica por dimensión (incluidos los valores
    ``SIN CLASIFICAR``), el resumen de la migración y las AFP detectadas.

    Args:
        src_bytes: Contenido binario de la planilla madre de licencias.
        dim_est_bytes: Contenido binario del maestro ``Establecimientos.xlsx``.
        log_callback: Callable opcional ``(str) -> None`` que recibe cada
            línea del log de progreso (la UI lo muestra en pantalla).
        pbit_data: Contenido binario opcional de la plantilla Power BI
            (``.pbit``) para incluirla en la salida.

    Returns:
        Dict ``{nombre_archivo: bytes}`` con:

        * ``01_Dim_Funcionario.xlsx``
        * ``02_Dim_Establecimiento.xlsx`` (incluye los establecimientos
          NUEVOS detectados, con tipo "Otro")
        * ``03_Dim_AFP.xlsx``
        * ``04_Hechos_Licencias.xlsx`` (hechos migrados + 40 filas en blanco)
        * ``05_Hechos_Descuentos.xlsx``
        * ``Dashboard_Licencias.pbit`` (solo si se pasó ``pbit_data``)
        * ``SLEP_files.zip`` con todo lo anterior comprimido
    """

    def log(msg):
        if log_callback:
            log_callback(msg)

    funcionarios, _establecimientos_raw, hojas_hechos = leer_fuente(src_bytes)
    dim_establecimientos = leer_dim_establecimiento(dim_est_bytes)

    # Logs iniciales 
    log("\n\nLeyendo planilla madre...")
    log(f"  {len(funcionarios)} funcionarios en DATOS")
    log(f"  {len(_establecimientos_raw)} establecimientos/unidades en catálogo inicial")

    # Folios repetidos
    folios = []
    for name, ws, hrow, headers in hojas_hechos:
        def _idx(*names):
            for n in names:
                if n in headers:
                    return headers.index(n)
            return None
        idx_folio = _idx("Folio licencia", "Folio Minsal")
        if idx_folio is None:
            continue
        for row in ws.iter_rows(min_row=hrow + 1, values_only=True):
            folio = row[idx_folio]
            if folio is not None and str(folio).strip():
                folios.append(str(folio).strip())

    folio_counts = Counter(folios)
    repetidos = {f: c for f, c in folio_counts.items() if c > 1}
    if repetidos:
        total_rep = sum(repetidos.values())
        log(f"  ▸ {len(repetidos)} folios repetidos ({total_rep} registros totales)")
    else:
        log("  ▸ 0 folios repetidos")
    log("\n")

    catalogo_norm = {norm(e["establecimiento"]): e["establecimiento"] for e in dim_establecimientos}
    catalogo_patterns = RE_ESTABLECIMIENTO

    (tipo_canon, tipo_clasif), (institucion_canon, institucion_clasif), (resolucion_canon, resolucion_clasif) = generar_listas_canonicas(hojas_hechos)

    # Listas canónicas
    log("Generando listas canónicas desde datos históricos...")
    log("")
    log("=== Tipo Licencia ===")
    for canon in tipo_canon:
        vals = tipo_clasif.get(canon, [])
        if vals:
            log(f"  '{canon}' <- {sorted(vals)}")
    sin = tipo_clasif.get("_sin_clasificar", [])
    if sin:
        log(f"  SIN CLASIFICAR: {sorted(sin)}")
    log("\n")

    log("=== Institución Salud ===")
    for canon in institucion_canon:
        vals = institucion_clasif.get(canon, [])
        if vals:
            log(f"  '{canon}' <- {sorted(vals)}")
    sin = institucion_clasif.get("_sin_clasificar", [])
    if sin:
        log(f"  SIN CLASIFICAR: {sorted(sin)}")
    log("\n")

    log("=== Resolución Médica ===")
    for canon in resolucion_canon:
        vals = resolucion_clasif.get(canon, [])
        if vals:
            log(f"  '{canon}' <- {sorted(vals)}")
    sin = resolucion_clasif.get("_sin_clasificar", [])
    if sin:
        log(f"  SIN CLASIFICAR: {sorted(sin)}")
    log("\n")

    # Migración
    log("Migrando hechos históricos...")
    hechos, nuevos_estab = migrar_hechos(
        hojas_hechos, funcionarios, catalogo_norm, catalogo_patterns,
        tipo_canon, institucion_canon, resolucion_canon,
    )

    log(f"  {len(hechos)} filas de hechos migradas desde {len(hojas_hechos)} hojas")
    if nuevos_estab:
        log(f"  {len(nuevos_estab)} establecimientos que requieren revisión: {', '.join(nuevos_estab.keys())}")

    descuentos = migrar_descuentos(hojas_hechos)

    for ne in nuevos_estab:
        dim_establecimientos.append({
            "tipo": "Otro",
            "establecimiento": ne,
            "comuna": None,
            "direccion": None,
            "telefono": None,
            "sitio_web": None,
        })

    dim_afp = construir_dim_afp(hojas_hechos)
    log(f"  {len(dim_afp)} combinaciones AFP+Tasa detectadas")

    out = {
        "01_Dim_Funcionario.xlsx": escribir_dim_funcionario(funcionarios),
        "02_Dim_Establecimiento.xlsx": escribir_dim_establecimiento(dim_establecimientos),
        "03_Dim_AFP.xlsx": escribir_dim_afp(dim_afp),
        "04_Hechos_Licencias.xlsx": escribir_hechos(
            funcionarios, dim_afp, hechos, tipo_canon, institucion_canon, resolucion_canon,
            descuentos=descuentos,
        ),
        "05_Hechos_Descuentos.xlsx": escribir_hechos_descuentos(descuentos),
    }
    if pbit_data is not None:
        out["Dashboard_Licencias.pbit"] = pbit_data

    n_hechos = len(hechos)
    log(f"04_Hechos_Licencias.xlsx generado con {n_hechos + 40} filas ({n_hechos} migradas + filas en blanco).")

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, data in out.items():
            if name == "SLEP_files.zip":
                continue
            zf.writestr(name, data)
    zip_buf.seek(0)
    out["SLEP_files.zip"] = zip_buf.read()
    return out
